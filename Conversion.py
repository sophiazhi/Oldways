# import xlrd # read xlsx or xls files
# from xlwt import Workbook # write xls files
from openpyxl import Workbook, load_workbook # read and modify Excel 2010 files


loc_order_report = ("StoreOrdersExport_Example.xlsx")
loc_sku_map = ("SKU_class_item.xlsx")

wb_order_report = load_workbook(loc_order_report, data_only=False)
order_sheet = wb_order_report.worksheets[0] #make a new sheet object

wb_sku_map = load_workbook(loc_sku_map)
sku_sheet = wb_sku_map.worksheets[0]

num_rows=order_sheet.max_row #number of rows
num_cols=order_sheet.max_column #number of columns

# print(num_rows, num_cols)
# print(order_sheet['A1'].value)
'''
print(sheet.cell_value(1,1)) #get the specific text at row 1, col 1

for i in range(num_of_cols):
	print(sheet.cell_value(0,i)) #print all the column names 
'''

wb_new = Workbook()
# sheet1 = wb_new.add_sheet("Modified Workbook") # xlwt 
ws1 = wb_new.active
ws1.title = "Sales Receipts"


# write column names
column_names = ['Customer', 'Date', 'Ref No.', 'Class', 'Payment method', 'Memo', 'Item', 'Quantity', 'Amount', 'Amount of Sales Receipt', 'Amount of transaction', 'Amount Deposited', 'Date deposited to CTC', 'Template Name']

for i in range(len(column_names)):
	ws1.cell(0+1, i+1, column_names[i])

for row in range(2, num_rows+1):
 	# for each number in total num. unique items, column S = column 19
 	ws1.cell(row, 1, "PRODUCTS")

 	date = order_sheet['C'+str(row)].value
 	# format date
 	ws1.cell(row, 2, date)
 	
 	customer_name = order_sheet['N'+str(row)].value
 	ws1.cell(row, 3, customer_name)
 	
 	sku_row = 0
 	for j in range(1,sku_sheet.max_row+1):
 		if sku_sheet['C'+str(j)].value == order_sheet['BD'+str(row)].value:
 			sku_row = j

 	if sku_row == 0:
 		# leave class blank because we don't know the sku
 		ws1.cell(row, 4, "")
 	else:
 		sku_class = sku_sheet['A'+str(sku_row)].value
 		ws1.cell(row, 4, sku_class)
 	
 	payment_method = order_sheet['R'+str(row)].value
 	ws1.cell(row, 5, payment_method)

 	memo = order_sheet['A'+str(row)].value
 	ws1.cell(row, 6, memo)

 	if sku_row == 0:
 		# leave item blank because we don't know the sku
 		ws1.cell(row, 7, "")
 	else:
 		sku_item = sku_sheet['B'+str(sku_row)].value
 		ws1.cell(row, 7, sku_item)
 	
 	# quantity unknown
 	
 	# amount of sales received unknown
 	
 	# amount of transaction = cost of total order
 	transaction = order_sheet['K'+str(row)].value
 	ws1.cell(row, 11, transaction)
 	
 	# amount deposited blank
 	ws1.cell(row, 12, "")
 	
 	# column M blank
 	ws1.cell(row, 13, "")
 	
 	ws1.cell(row, 14, "Custom Sales Receipt")



wb_new.save("Sales Receipts.xlsx")